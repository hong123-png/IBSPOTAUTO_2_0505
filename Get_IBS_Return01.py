import requests
import time
import pandas as pd
import os
import openpyxl
import datetime
from globalFunctions import log_msg
import sys
from fileArrangeCenter import fileArrangeCenterBetweenAB

processId = ''
now = datetime.datetime.now()
now = now.strftime("%Y-%m-%d")
log_file = os.path.join(os.path.dirname(os.path.abspath(__file__))+'\\log\\', 'stepA'+processId+now+'.log')  # 日志文件路径
print('====================== stepA'+processId+'程序开始运行 ======================')
log_msg(log_file, "stepA"+processId+"程序开始运行")

# 加载 Excel 文件
def load_excel(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return workbook, sheet
    except Exception as e:
        print(f"加载 Excel 文件失败: {e}")
        log_msg(log_file, f"加载 Excel 文件失败: {e}")
        return None, None

# 保存 Excel 文件
def save_excel(workbook, file_path):
    try:
        workbook.save(file_path)
        print(f"Excel 文件已保存: {file_path}")
        log_msg(log_file, f"Excel 文件已保存: {file_path},下一步是rename这个文件加完成标志")
    except Exception as e:
        print(f"保存 Excel 文件失败: {e}")
        log_msg(log_file, f"保存 Excel 文件失败: {e}")

        
# 搜索 IBSpot 并更新 Excel
def search_ibspot(file_path):
    # 加载 Excel 文件
    workbook, sheet = load_excel(file_path)
    if not workbook or not sheet:
        return

    # 获取最后一列的索引
    last_column = sheet.max_column  # 获取当前最后一列的索引

    # 检查最后一列是否为空
    while sheet.cell(row=1, column=last_column).value is None and last_column > 1:
        last_column -= 1

    # 结果插入到最后一列的下一列
    result_column = last_column + 1

    line = 1
    # 遍历每一行（从第二行开始）
    errorCount = 0
    for row in sheet.iter_rows(min_row=2, values_only=False):
        product_sku = row[1].value  # 第二列是产品SKU
        if not product_sku:
            continue

        print("操作第{}行的数据".format(line))
        log_msg(log_file, f"操作第{line}行的数据")
        search_url = f"http://35.222.94.60/api/product/check/{product_sku}"  # 直接拼接标题，不编码
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7",
            "Referer": "http://35.222.94.60/",
            "Connection": "keep-alive"
        }

        try:
            # 发送 HTTP 请求
            response = requests.get(search_url, headers=headers,proxies={"http": None, "https": None}) # 不使用代理
            status_code = response.status_code
            data = response.json()

            print(f"URL: {search_url}")
            print(f"Status Code: {status_code}")    
            log_msg(log_file, f"Status Code: {status_code}||URL: {search_url}")

            if status_code == 200:
                # 提取产品情况
                result_count = data["data"]["exist"]  # 提取 exist 字段的值

                if result_count is False:
                    sheet.cell(row=row[0].row, column=result_column).value = 0  # 插入到最后一列
                    print(f"修改结果: 0 (产品不存在)")
                    log_msg(log_file, f"修改结果: 0 (产品不存在)")
                elif result_count is True:
                    sheet.cell(row=row[0].row, column=result_column).value = 1  # 插入到最后一列
                    print(f"修改结果: 1 (产品存在)")
                    log_msg(log_file, f"修改结果: 1 (产品存在)")

                # 检查是否修改成功
                if sheet.cell(row=row[0].row, column=result_column).value in [0, 1]:
                    print("修改成功: 值已更新为", sheet.cell(row=row[0].row, column=result_column).value)
                    log_msg(log_file, f"修改成功: 值已更新为 {sheet.cell(row=row[0].row, column=result_column).value}")
                else:
                    print("修改失败: 值未更新")
                    log_msg(log_file, "修改失败: 值未更新")
            else:
                sheet.cell(row=row[0].row, column=result_column).value = status_code  # 搜索失败，插入状态码
                print(f"修改结果: {status_code} (状态码: {status_code})")
                log_msg(log_file, f"修改结果: {status_code} (状态码: {status_code})")

                # 检查是否修改成功
                if sheet.cell(row=row[0].row, column=result_column).value == status_code:
                    print("修改成功: 值已更新为", status_code)
                else:
                    print("修改失败: 值未更新")
        except Exception as e:
            errorCount += 1
            if errorCount > 7:
                print("发生异常次数超过7次,程序将退出。")
                log_msg(log_file, "发生异常次数超过7次,程序将退出。")
                break
            sheet.cell(row=row[0].row, column=result_column).value = -1  # 发生异常，插入-1
            print(f"修改结果: -1 (异常: {e})")
            log_msg(log_file, f"修改结果: -1 (异常: {e})")

            # 检查是否修改成功
            if sheet.cell(row=row[0].row, column=result_column).value == -1:
                print("修改成功: 值已更新为 -1")
                log_msg(log_file, "修改成功: 值已更新为 -1")
            else:
                print("修改失败: 值未更新")
                log_msg(log_file, "修改失败: 值未更新")

        time.sleep(1)  # 增加1秒延迟，避免频繁请求
        line = line + 1

    # 处理 file_path， 找到文件名，在前面加上 IBA_DONE_，并保存
    # file_name = os.path.basename(file_path)
    # new_file_name = f"IBA_DONE_{file_name}"
    # new_file_path = os.path.join(os.path.dirname(file_path), new_file_name)
    # 保存 Excel 文件
    save_excel(workbook, file_path)
    

# 将 Excel 文件另存为 CSV 格式
def save_as_csv(excel_file, csv_file):
    try:
        # 使用 pandas 读取 Excel 文件
        df = pd.read_excel(excel_file)

        # 过滤掉最后一列不为 0 的行
        df_filtered = df[df.iloc[:, -1] == 0]
        # 移除 '$' 符号和逗号，并将列转换为浮点数类型
        # df_filtered['cost_price'] = df_filtered['cost_price'].str.replace('$', '', regex=False).str.replace(',', '', regex=False).astype(float)
        # 过滤 'cost_price' 小于 150 的行
        # df_filtered = df_filtered[df_filtered['cost_price'] < 150]
        # 检查过滤后的 DataFrame 是否为空
        if not df_filtered.empty:
            # 将过滤后的 DataFrame 保存为 CSV 格式
            df_filtered.to_csv(csv_file, index=False, encoding='utf-8')
            print(f"Excel 文件已成功保存为 CSV 格式: {csv_file}")
            log_msg(log_file, f"Excel 文件已成功保存为 CSV 格式: {csv_file}")
        else:
            print(f"过滤后没有符合条件的行 (最后一列为 0)，CSV 文件未保存。")
            log_msg(log_file, f"过滤后没有符合条件的行 (最后一列为 0)，CSV 文件未保存。")
    except Exception as e:
        print(f"保存 CSV 文件失败: {e}")
        log_msg(log_file, f"保存 CSV 文件失败: {e}")


def process_files(directory):
    lock_fileA = os.path.join(directory, "LOCK1.txt")  # 构建完整的文件路径
    if os.path.exists(lock_fileA):
        log_msg(log_file, f"目录 '{directory}' 下的 LOCK1.txt 文件已存在，程序将退出。")
        return
    else:
        try:    
            # 创建文件
            with open(lock_fileA, "w") as f:
                f.write("Get_IBS_Return01 is working.")
                log_msg(log_file, f"创建了目录 '{directory}' 下的 LOCK1.txt 文件。")

            # file_for_count = 0
            # 遍历目录中的所有文件
            for file_name in os.listdir(directory):
                # 如果文件名以 '.xlsx' 结尾，跳过该文件
                if file_name.startswith('IBA_DONE_') or not file_name.endswith('.xlsx'):
                    continue

                # file_for_count += 1
                file_path = os.path.join(directory, file_name)
                print(f"正在处理文件: {file_path}")
                log_msg(log_file, f"正在处理文件: {file_path}")

                # 处理文件
                search_ibspot(file_path)
                # if file_for_count % 2 == 0:
                #     saved_csv_path = os.path.join(os.path.dirname(file_path)+"_a1", f"{os.path.splitext(os.path.basename(file_path))[0]}.csv")
                # else:
                #     saved_csv_path = f"{os.path.splitext(file_path)[0]}.csv"
                saved_csv_path = fileArrangeCenterBetweenAB()
                if saved_csv_path:
                    saved_csv_path = os.path.join(saved_csv_path, f"{os.path.splitext(os.path.basename(file_path))[0]}.csv")
                    # 保存为 CSV 格式
                    save_as_csv(file_path, saved_csv_path)
                    log_msg(log_file, f"文件 {file_path} 已保存为 CSV 格式: {saved_csv_path}")
                else:
                    print(f"文件夹 {saved_csv_path} 不存在， ArrangeCenter保存文件失败。")
                    log_msg(log_file, f"文件夹 {saved_csv_path} 不存在， ArrangeCenter保存文件失败。")
                # save_as_csv(file_path, saved_csv_path)

                renamed_file_path = os.path.join(os.path.dirname(file_path), f"IBA_DONE_{os.path.basename(file_path)}")
                if os.path.exists(renamed_file_path):
                    os.remove(renamed_file_path)
                os.rename(file_path, os.path.join(os.path.dirname(file_path), f"IBA_DONE_{os.path.basename(file_path)}"))
        finally:
            # 删除文件
            if os.path.exists(lock_fileA):
                os.remove(lock_fileA)
                print(f"成功删除了目录 '{directory}' 下的 LOCK1.txt 文件。")
                log_msg(log_file, f"成功删除了目录 '{directory}' 下的 LOCK1.txt 文件。")
            log_msg(log_file, "StepA"+processId+"程序执行完毕")
            print("====================== StepA"+processId+"程序执行完毕 ======================")
# 主函数
if __name__ == "__main__":
    directory = "untreated"+processId  # 替换为你的文件夹路径
    process_files(directory)
    sys.exit(0)
    